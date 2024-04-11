from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from itertools import product

import threading

_thread_lock = threading.Lock()


class Person:
	def __init__(self):
		self.last_name: str = ''
		self.first_name: str = ''
		self.middle_name: str = ''

	@property
	def full_name(self) -> str:
		if self.middle_name:
			full_name = f'{self.last_name.capitalize()} {self.first_name.capitalize()} {self.middle_name.capitalize()}'
		else:
			full_name = f'{self.last_name.capitalize()} {self.first_name.capitalize()}'

		return full_name

	@classmethod
	def from_dict(cls, person_dict: dict) -> 'Person':
		if not all(person_dict.get(key) for key in ['Фамилия', 'Имя']):
			raise ValueError('Incorrect person information - No first name or last name')

		cls_attrs = {
			'Фамилия': 'last_name',
			'Имя': 'first_name',
			'Отчество': 'middle_name',
		}
		person = cls()
		for dict_attr, cls_attr in product(person_dict, cls_attrs):
			if (cls_attr.lower() in dict_attr.lower() or dict_attr.lower() in cls_attr.lower()) and person_dict.get(dict_attr):
				raw_attr_value = person_dict.get(dict_attr)
				attr_value = str(raw_attr_value).strip()
				setattr(person, cls_attrs.get(cls_attr), attr_value)
		return person


def get_debtors_list(input_excel_file: str) -> list[Person]:
	persons_table_workbook = load_workbook(filename=input_excel_file)
	persons_table_sheet = persons_table_workbook.active

	header_row = [cell.value for cell in persons_table_sheet[1]]

	person_dict_template = {
		'Фамилия': '',
		'Имя': '',
		'Отчество': '',
	}
	persons_list = []
	for row in persons_table_sheet.iter_rows(min_row=2, values_only=True):
		person_dict = person_dict_template.copy()
		for cell_value, col_name in zip(row, header_row):
			if col_name:
				try:
					key = next(key for key in person_dict if col_name in key or key in col_name)
				except StopIteration:
					raise ValueError('Incorrect columns in input excel file')

				person_dict[key] = cell_value if cell_value else ''

		if all(person_dict.get(key) for key in ['Фамилия', 'Имя']):
			person = Person.from_dict(person_dict=person_dict)
			persons_list.append(person)

	if not persons_list:
		raise ValueError('Incorrect input - No person information')

	return persons_list


def output_check_result(output_file: str, check_result: dict[str: str]) -> None:
	with _thread_lock:
		try:
			wb: Workbook = load_workbook(filename=output_file)
			ws = wb.active
			header = False
		except FileNotFoundError:
			wb = Workbook()
			ws = wb.create_sheet(title='Банкротство')
			header = True

		for row in dataframe_to_rows(df=pd.json_normalize(data=check_result), index=False, header=header):
			ws.append(row)

		if header:
			for sheet_name in wb.sheetnames:
				sheet = wb[sheet_name]
				if sheet.max_row == 1 and sheet.max_column == 1:
					wb.remove(worksheet=sheet)

		wb.save(filename=output_file)


if __name__ == '__main__':
	for pers in get_debtors_list(input_excel_file='input/persons_list.xlsx'):
		print(pers.full_name)
